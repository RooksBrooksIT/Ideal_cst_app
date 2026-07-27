import openpyxl
import requests
import json
import re
import random
import string
import sys

PROJECT_ID = 'cst-pratap-demo'
BASE_URL = f'https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents'

def generate_doc_id(length=20):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

def make_firestore_value(val):
    if val is None:
        return {'nullValue': None}
    elif isinstance(val, bool):
        return {'booleanValue': val}
    elif isinstance(val, int):
        return {'integerValue': str(val)}
    elif isinstance(val, float):
        if val.is_integer():
            return {'integerValue': str(int(val))}
        return {'doubleValue': val}
    elif isinstance(val, str):
        return {'stringValue': val}
    elif isinstance(val, list):
        return {'arrayValue': {'values': [make_firestore_value(x) for x in val]}}
    elif isinstance(val, dict):
        return {'mapValue': {'fields': {k: make_firestore_value(v) for k, v in val.items()}}}
    else:
        return {'stringValue': str(val)}

def create_firestore_doc(collection_path, doc_id, fields):
    url = f'{BASE_URL}/{collection_path}?documentId={doc_id}'
    body = {'fields': {k: make_firestore_value(v) for k, v in fields.items()}}
    res = requests.post(url, json=body)
    if res.status_code == 409: # Already exists, patch/update
        url_patch = f'{BASE_URL}/{collection_path}/{doc_id}'
        res = requests.patch(url_patch, json=body)
    return res

# 1. Fetch Firestore Sites
print('Fetching Firestore Sites...', flush=True)
res_sites = requests.get(f'{BASE_URL}/Site?pageSize=300')
sites_docs = res_sites.json().get('documents', [])
site_by_name = {}
for s in sites_docs:
    f = s['fields']
    doc_id = s['name'].split('/')[-1]
    s_id = f.get('siteId', {}).get('stringValue', '')
    s_name = f.get('siteName', {}).get('stringValue', '')
    clean_n = s_name.lower().replace(' ', '').replace('-', '').replace('.', '')
    site_by_name[clean_n] = {'doc_id': doc_id, 'siteId': s_id, 'siteName': s_name}

# 2. Fetch Firestore Supervisors
print('Fetching Firestore Supervisors...', flush=True)
res_sups = requests.get(f'{BASE_URL}/supervisor?pageSize=300')
sups_docs = res_sups.json().get('documents', [])
sup_map = {}
for sp in sups_docs:
    f = sp['fields']
    doc_id = sp['name'].split('/')[-1]
    sp_id = f.get('SupervisorId', {}).get('stringValue', doc_id)
    u_name = f.get('UserName', {}).get('stringValue', '')
    f_name = f.get('FullName', {}).get('stringValue', '')
    entry = {'doc_id': doc_id, 'supervisorId': sp_id, 'userName': u_name, 'fullName': f_name}
    sup_map[u_name.lower()] = entry
    sup_map[f_name.lower()] = entry
    clean_f = f_name.lower().replace('mr.', '').replace('mr ', '').strip()
    sup_map[clean_f] = entry

sup_map['krishna kumar'] = {'doc_id': 'SP021_KK', 'supervisorId': 'SP021_KK', 'userName': 'KK', 'fullName': 'KrishnaKumar'}

# 3. Fetch SiteSupervisorMap
print('Fetching SiteSupervisorMap...', flush=True)
res_ssm = requests.get(f'{BASE_URL}/siteSupervisorMap?pageSize=300')
ssm_docs = res_ssm.json().get('documents', [])
site_sup_map = {}
for ssm in ssm_docs:
    f = ssm['fields']
    doc_id = ssm['name'].split('/')[-1]
    s_id = f.get('siteId', {}).get('stringValue', doc_id)
    sup_user = f.get('supervisor', {}).get('stringValue', '')
    site_sup_map[doc_id.lower()] = {'map_doc_id': doc_id, 'siteId': s_id, 'supervisor': sup_user}

# 4. Fetch SubContractors
print('Fetching SubContractors...', flush=True)
res_sc = requests.get(f'{BASE_URL}/sub_contractors?pageSize=300')
sc_docs = res_sc.json().get('documents', [])
sc_map = {}
for sc in sc_docs:
    f = sc['fields']
    doc_id = sc['name'].split('/')[-1]
    sc_name = f.get('name', {}).get('stringValue', '')
    sc_map[sc_name.lower().strip()] = {
        'doc_id': doc_id,
        'contractorId': f.get('contractorId', {}).get('stringValue', doc_id),
        'salaryRate': float(f.get('salaryRate', {}).get('integerValue') or f.get('salaryRate', {}).get('doubleValue') or 850),
        'overtimeRate': float(f.get('overtimeRate', {}).get('integerValue') or f.get('overtimeRate', {}).get('doubleValue') or 100),
        'category': f.get('category', {}).get('stringValue', ''),
        'assignedSiteIds': [v.get('stringValue') for v in f.get('assignedSiteIds', {}).get('arrayValue', {}).get('values', [])] if 'assignedSiteIds' in f else []
    }

CT_MAP = {
    'MH': 'M.Helper',
    'M': 'Mason',
    'CM': 'Concrete Mason',
    'CH': 'CH',
    'CON.L': 'Con.L',
    'CON. L': 'Con.L',
    'TMA': 'Tiles Mason Assistant',
    'PR': 'Painter',
    'PLU': 'Plumber',
    'ELE': 'Electrician',
    'WAT': 'Watchman',
    'WATCHMAN': 'Watchman',
    'WL': 'Welder',
    'MODULAR KITCHEN': 'Modular Kitchen',
    'CUPBOARD': 'Cupboard',
    'HITACHI': 'Hitachi',
    'TIPPER': 'Tipper',
    'OTHERS': 'Others'
}

LT_MAP = {
    'SC': 'Sub Contractor',
    'DW': 'Daily Wage'
}

# Parse Excel
print('Parsing Excel file...', flush=True)
wb = openpyxl.load_workbook('Site Labour Details - 24.07.2026-2 (1).xlsx', data_only=True)
sheet = wb['mar']

current_coord = None
current_site = None
current_supervisor = None

parsed_rows = []

for r in range(3, sheet.max_row + 1):
    s_no = sheet.cell(r, 1).value
    coord = sheet.cell(r, 2).value
    site = sheet.cell(r, 3).value
    supervisor = sheet.cell(r, 4).value
    ct = sheet.cell(r, 5).value
    lt = sheet.cell(r, 6).value
    sub_contractor = sheet.cell(r, 7).value
    nos = sheet.cell(r, 8).value
    ot_details = sheet.cell(r, 10).value

    if s_no is None and coord is None and site is None and supervisor is None and ct is None and lt is None and sub_contractor is None:
        continue

    if coord is not None and str(coord).strip():
        current_coord = str(coord).strip()
    
    if site is not None and str(site).strip():
        s_str = str(site).strip()
        if 'Bejansingh Projects' in s_str or 'NOH(P)' in s_str:
            continue
        current_site = s_str

    if supervisor is not None and str(supervisor).strip():
        current_supervisor = str(supervisor).strip()

    # Mandatory field validation
    if not current_coord or not current_site or not current_supervisor or not ct or not lt or not sub_contractor or nos is None:
        continue

    parsed_rows.append({
        'row': r,
        'coord': current_coord,
        'site': current_site,
        'supervisor': current_supervisor,
        'ct': str(ct).strip(),
        'lt': str(lt).strip(),
        'sub': str(sub_contractor).strip(),
        'nos': int(nos),
        'ot': str(ot_details).strip() if ot_details else None
    })

print(f'Total valid candidate rows for 24.07.2026: {len(parsed_rows)}', flush=True)

# Group candidate rows by siteId
site_entries = {}

for item in parsed_rows:
    site_str = item['site']
    sup_str = item['supervisor']
    
    clean_s = site_str.lower().replace(' ', '').replace('-', '').replace('.', '')
    matched_site = site_by_name.get(clean_s)
    if not matched_site:
        for k, v in site_by_name.items():
            if clean_s in k or k in clean_s:
                matched_site = v
                break
    
    site_doc_id = matched_site['doc_id'] if matched_site else site_str

    clean_sup = sup_str.lower().replace('mr.', '').replace('mr ', '').strip()
    first_sup = clean_sup.split(',')[0].strip()
    matched_sup = sup_map.get(first_sup) or sup_map.get(clean_sup)
    
    sup_id_val = matched_sup['supervisorId'] if matched_sup else 'SP001'
    sup_name_val = matched_sup['userName'] if matched_sup else sup_str

    map_key = f"{site_doc_id}_{sup_id_val}".lower()
    full_site_id = f"{site_doc_id}_{sup_id_val}"
    
    if map_key in site_sup_map:
        full_site_id = site_sup_map[map_key]['map_doc_id']
    
    item['resolved_site_id'] = full_site_id
    item['resolved_site_name'] = site_doc_id
    item['resolved_supervisor_id'] = sup_id_val
    item['resolved_supervisor_name'] = sup_name_val

    if full_site_id not in site_entries:
        site_entries[full_site_id] = {
            'siteId': full_site_id,
            'siteName': site_doc_id,
            'supervisorId': sup_id_val,
            'supervisorName': sup_name_val,
            'coordinatorName': item['coord'],
            'date': '2026-07-24',
            'notes': '',
            'weather': '',
            'items': []
        }
    site_entries[full_site_id]['items'].append(item)

print(f'Total distinct Site-Supervisor entries to write: {len(site_entries)}', flush=True)

total_workers_created = 0
site_count = 0

for site_id, s_data in site_entries.items():
    site_count += 1
    parent_doc_id = f"{site_id}_2026-07-24"

    worker_docs = []
    
    for item in s_data['items']:
        ct_upper = item['ct'].upper()
        category = CT_MAP.get(ct_upper, item['ct'])
        lt_upper = item['lt'].upper()
        labour_type = LT_MAP.get(lt_upper, item['lt'])
        sub_name = item['sub']
        nos = item['nos']
        ot_text = item['ot']

        in_time = "09:00 AM"
        out_time = "05:00 PM"
        attendance_type = "Full Day"
        day_value = 1.0
        hours_worked = 8.0
        default_hours = 8.0
        ot_hours_str = "0 Hours"
        ot_hours_num = 0.0
        remarks = ""

        if ot_text:
            ot_lower = ot_text.lower()
            if '1.00 p.m. out' in ot_lower or '11.00 a.m. out' in ot_lower:
                out_time = "01:00 PM" if '1.00' in ot_lower else "11:00 AM"
                attendance_type = "Half Day"
                day_value = 0.5
                hours_worked = 4.0
            elif '11.00 a.m. to 1.00 p.m.' in ot_lower:
                in_time = "11:00 AM"
                out_time = "01:00 PM"
                attendance_type = "Half Day"
                day_value = 0.5
                hours_worked = 2.0
            elif '2.00 p.m. to 6.00 p.m.' in ot_lower:
                in_time = "02:00 PM"
                out_time = "06:00 PM"
                attendance_type = "Half Day"
                day_value = 0.5
                hours_worked = 4.0
            elif 'curing' in ot_lower:
                remarks = ot_text
                if '5 hrs' in ot_lower:
                    ot_hours_str = "5 Hours"
                    ot_hours_num = 5.0
                elif '3 hrs' in ot_lower:
                    ot_hours_str = "3 Hours"
                    ot_hours_num = 3.0
            elif 'ot 6.00 p.m. to 8.30 p.m.' in ot_lower:
                out_time = "08:30 PM"
                ot_hours_str = "2.5 Hours"
                ot_hours_num = 2.5
                hours_worked = 10.5
            elif '6.00 a.m. to' in ot_lower:
                in_time = "06:00 AM"
                ot_hours_str = "3 Hours"
                ot_hours_num = 3.0
                hours_worked = 11.0
            elif '7.00 a.m. to' in ot_lower:
                in_time = "07:00 AM"
                ot_hours_str = "2 Hours"
                ot_hours_num = 2.0
                hours_worked = 10.0
            elif '10.30 a.m. to' in ot_lower:
                in_time = "10:30 AM"
            elif '12.00 p.m. to' in ot_lower:
                in_time = "12:00 PM"
            elif '2.30 p.m. to' in ot_lower:
                in_time = "02:30 PM"

        sc_info = sc_map.get(sub_name.lower().strip())
        contractor_id = sc_info['doc_id'] if sc_info else generate_doc_id()
        basic_salary = sc_info['salaryRate'] if sc_info else 850.0
        overtime_rate = sc_info['overtimeRate'] if sc_info else 100.0
        overtime_amount = ot_hours_num * overtime_rate
        total_salary = (basic_salary * day_value) + overtime_amount

        worker_names = []
        if nos == 1:
            worker_names.append(sub_name)
        else:
            for w in range(1, nos + 1):
                worker_names.append(f"{sub_name} W{w}")

        for w_idx, w_name in enumerate(worker_names):
            worker_doc_id = generate_doc_id()
            w_entry = {
                'workerId': worker_doc_id,
                'workerName': w_name,
                'contractorName': sub_name,
                'contractorId': contractor_id,
                'isContractor': True if nos == 1 or w_idx == 0 else False,
                'category': category,
                'labourType': labour_type,
                'attendanceType': attendance_type,
                'inTime': in_time,
                'outTime': out_time,
                'otHours': ot_hours_str,
                'dayValue': day_value,
                'defaultHours': default_hours,
                'hoursWorked': hours_worked,
                'overtimeHours': ot_hours_num,
                'overtimeRate': overtime_rate,
                'overtimeAmount': overtime_amount,
                'basicSalary': basic_salary,
                'totalSalary': total_salary,
                'mealsCount': 0,
                'mealsAmount': 0,
                'busCount': 0,
                'busAmount': 0,
                'remarks': remarks,
                'siteId': s_data['siteId'],
                'siteName': s_data['siteName'],
                'supervisorId': s_data['supervisorId'],
                'supervisorName': s_data['supervisorName'],
                'coordinatorName': s_data['coordinatorName'],
                'date': '2026-07-24',
                'mobile': '0000000000'
            }
            worker_docs.append(w_entry)
            total_workers_created += 1

    parent_fields = {
        'siteId': s_data['siteId'],
        'siteName': s_data['siteName'],
        'supervisorId': s_data['supervisorId'],
        'supervisorName': s_data['supervisorName'],
        'coordinatorName': s_data['coordinatorName'],
        'date': '2026-07-24',
        'notes': '',
        'weather': '',
        'totalWorkers': len(worker_docs),
        'summary': {
            'fullDay': sum(1 for w in worker_docs if w['attendanceType'] == 'Full Day'),
            'halfDay': sum(1 for w in worker_docs if w['attendanceType'] == 'Half Day'),
            'absent': 0,
            'leave': 0,
            'earlyOut': 0,
            'effectiveLabourCount': sum(w['dayValue'] for w in worker_docs),
            'totalOTHours': sum(w['overtimeHours'] for w in worker_docs)
        }
    }

    create_firestore_doc('daily_labour_entries', parent_doc_id, parent_fields)

    for w in worker_docs:
        w_id = w['workerId']
        create_firestore_doc(f'daily_labour_entries/{parent_doc_id}/workers', w_id, w)
    
    print(f'[{site_count}/{len(site_entries)}] Wrote {parent_doc_id} with {len(worker_docs)} workers.', flush=True)

print(f'\nIMPORT COMPLETE!', flush=True)
print(f'Successfully imported 24.07.2026 data across {len(site_entries)} parent site documents and {total_workers_created} worker documents.', flush=True)
